import numpy as np
import matplotlib.pyplot as plt
import pydicom
import shutil
import random
import torch.nn as nn
import os
import SimpleITK as sitk
import ants
import cv2
import nibabel as nib
import imageio
from skimage.transform import resize
from scipy.ndimage import zoom
from collections import Counter
import openpyxl


############# 文件操作模块 #############
def copy_and_paste():
    # path = "../../../data/diag_data/"  # CT_CTA disease
    catch = "../../../data/catch"
    if not os.path.isdir(catch):
        os.makedirs(catch)

    f = open("p_test1.txt")  # test train！！  list  filter
    path_list = []
    for line in f.readlines():
        if "SE0" in line and line.split('/IM')[0] not in path_list:
            path_list.append(line.split('/IM')[0])
    path_list.sort()
    ii = 0
    for sub_path in path_list:
        input_files = os.listdir(sub_path)
        input_files.sort()
        input_files.sort(key=lambda x: (int(x.split('IM')[1])))
        sub_pathe2=sub_path.replace("SE0", "SE1")
        diag=sub_path.split("/")[4]
        sub_path_out1= sub_path.replace(diag, "diag1a_data")
        sub_path_out2 = sub_path_out1.replace("SE0", "SE1")
        if not os.path.isdir(sub_path_out1):
            os.makedirs(sub_path_out1)

        if not os.path.isdir(sub_path_out2):
            os.makedirs(sub_path_out2)

        for j in range(len(input_files)):
            in_file_path1 = os.path.join(sub_path, input_files[j])
            in_file_path2 = os.path.join(sub_pathe2, input_files[j])
            out_file_path1 = os.path.join(sub_path_out1, input_files[j])
            out_file_path2 = os.path.join(sub_path_out2, input_files[j])
            shutil.copy(in_file_path1, out_file_path1)
            shutil.copy(in_file_path2, out_file_path2)

        ii=ii+1
        if ii%10==0:
            print('numbers:', ii)

#get real input list
def get_test_list():
    path="../../../data/p2_nii/"#neckk
    path_list=[]
    for root, dirs, files in os.walk(path, topdown=False):
        path_list.append(root)
    f = open("nor.txt", "w")
    for sub_path in path_list:
        data_files = os.listdir(sub_path)
        for j in range(len(data_files)):
            file_path=os.path.join(sub_path, data_files[j])
            if os.path.isdir(file_path):
                continue
            # file_path_list.append(file_path)
            f.writelines(file_path + "\n")
    f.close()

# 主函数
def nii_to_image():
    filepath="../../../data/ASL/"
    imgfile = "../../../data/ASL_img/"
    filenames = os.listdir(filepath)

    # 开始读取nii文件
    for f in filenames:
        img_path = os.path.join(filepath, f)
        img = nib.load(img_path)  # 读取nii
        img_fdata = img.get_fdata()
        # 去掉nii的后缀名
        fname = f.replace('.nii', '')
        img_f_path = os.path.join(imgfile, fname)

        # 创建nii对应的图像的文件夹
        if not os.path.isdir(img_f_path):
            os.makedirs(img_f_path)
        # if not os.path.exists(img_f_path):
        #     # 新建文件夹
        #     os.mkdir(img_f_path)
        #     # 开始转换为图像
        (x, y, z) = img.shape
        # z是图像的序列
        for i in range(z):
            # 选择哪个方向的切片都可以
            silce = img_fdata[:, :, i]
            # silce = (silce * 255.0).astype('uint8')
            # silce = silce.astype(np.uint8)
            imageio.imwrite(os.path.join(img_f_path, '{}.png'.format(i)), silce)

            # silce = silce.astype(np.uint8)
            # # silce = (silce * 255.0).astype('uint8')
            # silce = Image.fromarray(silce)
            # silce.save(os.path.join(img_f_path, '{}.png'.format(i)))

############# 自动构建mask #############
def to_windowdata(image,WC,WW):
    # image = (image + 1) * 0.5 * 4095
    # image[image == 0] = -2000
    # image=image-1024
    center = WC #40 400//60 300
    width = WW# 200
    try:
        win_min = (2 * center - width) / 2.0 + 0.5
        win_max = (2 * center + width) / 2.0 + 0.5
    except:
        # print(WC[0])
        # print(WW[0])
        center = WC[0]  # 40 400//60 300
        width = WW[0]  # 200
        win_min = (2 * center - width) / 2.0 + 0.5
        win_max = (2 * center + width) / 2.0 + 0.5
    dFactor = 255.0 / (win_max - win_min)
    image = image - win_min
    image = np.trunc(image * dFactor)
    image[image > 255] = 255
    image[image < 0] = 0
    image=image/255#np.uint8(image)
    # image = (image - 0.5)/0.5
    return image

def display():#file_path1,file_path2
    # file_pathE = '../output/Hd/img2/e1/4/ST0/SE0/IM11'
    file_path1 = '../../../data/p1_data/dis/dml/fb/DICOM3/PA3/ST0/SE0/IM6'

    dicom = sitk.ReadImage(file_path1)
    data1 = np.squeeze(sitk.GetArrayFromImage(dicom))
    data1 = data1+1024
    file_path2 = '../../../data/p1_data/dis/dml/fb/DICOM3/PA3/ST0/SE1/IM6'
    dicom = sitk.ReadImage(file_path2)
    data2 = np.squeeze(sitk.GetArrayFromImage(dicom))
    data2 = data2+1024
    data=data2-data1
    dsA = pydicom.dcmread(file_path1, force=True)
    # aa=ds.SeriesInstanceUID
    # ab=ds.FrameOfReferenceUID
    # ac=ds.StudyInstanceUID
    #
    # file_pathE = '../output/Hd1/img2/e40/4/ST0/SE1/IM11'
    # dsb = pydicom.dcmread(file_pathE, force=True)  # 读取头文件
    # aa2=dsb.SeriesInstanceUID
    # ab2=dsb.FrameOfReferenceUID
    # ac2=dsb.StudyInstanceUID

    # data = (ds.pixel_array).astype(np.int)
    # WC = ds.WindowCenter
    # WW = ds.WindowWidth
    # a = to_windowdata(data, WC, WW)
    image1=data1#sitk读取的数值比pydicom读取的数值小1024
    image1=image1/4095
    image1 = (image1- 0.5)/0.5
    image1[image1>1]=1
    image1[image1<=-1]=-1
    # kernel2 = np.ones((2, 2), np.uint8)
    # image1= cv2.dilate(image1, kernel2)  # [-1 1]


    image2=data1#sitk读取的数值比pydicom读取的数值小1024
    image2[image2<1037]=-2000#-2000->0
    image2[image2 > 1100] = -2000
    image2=image2/4095
    image2 = (image2- 0.5)/0.5
    image2[image2>-1]=1
    image2[image2<=-1]=0
    # kernel2 = np.ones((2, 0), np.uint8)
    # image2 = cv2.dilate(image2, kernel2)#[-1 1]


    image3=data2#sitk读取的数值比pydicom读取的数值小1024
    image3[image3<0]=0#-2000->0
    image3=image3/4095
    image3 = (image3 - 0.5)/0.5

    image4=data2#sitk读取的数值比pydicom读取的数值小1024
    image4=to_windowdata(image4, 120, 20)
    image5=image4*image2
    image5[image5<= 0] = -1
    kernel1 = np.ones((2, 2), np.uint8)
    image5 = cv2.dilate(image5, kernel1)  # [-1 1]
    # m=get_mask(image2)
    # image2=image2*m
    # image2[image2==0]=-1

    # image3=image1*image2
    # image3[image3 == 0] = -1

    # plt.close()
    plt.subplot(2, 2, 1)
    plt.imshow(image1, cmap='gray')  # ,vmin=0,vmax=255
    plt.subplot(2, 2, 2)
    plt.imshow(image2, cmap='gray')  # ,vmin=0,vmax=255
    plt.subplot(2, 2, 3)
    plt.imshow(image4, cmap='gray')  # ,vmin=0,vmax=255
    plt.subplot(2, 2, 4)
    plt.imshow(image5, cmap='gray')  # ,vmin=0,vmax=255
    plt.show()
    # cv2.imwrite("outpath"+'.png',((image2*0.5)+0.5)*255)

#dicom to nii
def dcm2nii_sitk(path_read, path_save):
    reader = sitk.ImageSeriesReader()
    seriesIDs = reader.GetGDCMSeriesIDs(path_read)
    N = len(seriesIDs)
    lens = np.zeros([N])
    for i in range(N):
        dicom_names = reader.GetGDCMSeriesFileNames(path_read, seriesIDs[i])
        lens[i] = len(dicom_names)
    N_MAX = np.argmax(lens)
    dicom_names = reader.GetGDCMSeriesFileNames(path_read, seriesIDs[N_MAX])
    reader.SetFileNames(dicom_names)
    image = reader.Execute()
    sitk.WriteImage(image, path_save)
    # shutil.rmtree(path_read)

def standardize_nii(path):
    image = sitk.ReadImage(path)
    size=image.GetSize()#GetSize (x,y,z)和numpy size(z,y,x)不同。
    direction = image.GetDirection()
    if size[0]!=800 or size[1]!=800:
        print(path)
        # 交换x和z轴
        new_size = (size[2], size[1], size[0])
        new_direction = (direction[6], direction[7], direction[8],
                         direction[3], direction[4], direction[5],
                         direction[0], direction[1], direction[2])
        # 创建一个新的图像，并设置新的大小和方向
        new_image = sitk.Image(new_size, image.GetPixelID())
        new_image.SetDirection(new_direction)

        # 将原始图像的像素值复制到新图像中
        new_image = sitk.Paste(new_image, image, image.GetSize(), [0, 0, 0])
        new_image_size = new_image.GetSize()
        if new_image_size[0] == 800 or new_image_size[1] == 800:
            print("corrected!" + "\n")
            sitk.WriteImage(new_image, path)  ## 保存修改后的图像
    else:
        new_image = image
    # aa=1

def standardize_nii_process():
    target_coordinate_system = 'RAS'
    path = "/root/autodl-tmp/yml/data/xy/breast/"  #
    path_list = []
    aug_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        if "0.nii.gz" in files:
            file_path = os.path.join(root, "0.nii.gz")
            path_list.append(file_path)

    for j in range(len(path_list)):
        standardize_nii(path_list[j])
        standardize_nii(path_list[j].replace("0.nii.gz", "1.nii.gz"))
    print("finished!")

def crop_by_mask():
    #保留动脉血流范围的图像；mask分别为动脉血流、钙化斑块、血栓、软斑块；如果只有动脉血流则保持不变。
    # path = "../../../data/p2_s/p2/dis/dmzyyh/DICOM0/PA1" #
    path = "../../../data/p2_s" #
    # path = "../../../data/p2_nii/hnnk/"#/jfj/p1/dis/dml/fb/DICOM0/PA0
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "2.nii.gz" in path:
                    path_list.append(path)
    i=0
    for se2output in path_list:
        se0output = se2output.replace("2.nii.gz", "0.nii.gz")
        read = sitk.ReadImage(se0output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array0 = sitk.GetArrayFromImage(read)
        img_array_w0 = to_windowdata(img_array0, 130, 10)
        img_array_w0 = np.where(img_array_w0 > 0, 1, 0)#支架与钙化斑块
        # img_array_w00=1-img_array_w0

        se1output = se2output.replace("2.nii.gz", "1.nii.gz")
        read = sitk.ReadImage(se1output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array1 = sitk.GetArrayFromImage(read)

        read = sitk.ReadImage(se2output, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
        img_array2 = sitk.GetArrayFromImage(read)
        #裁剪有效范围
        su=np.sum(img_array2,axis=(1,2))
        index=np.where(su>16)[0]#避免一些小点出现
        # index=np.nonzero(img_array2)
        st=min(index)
        en=max(index)+1

        # st=img_array1.shape[0]-img_array2.shape[0]#处理特殊情况
        # en=img_array1.shape[0]+1

        out2 = sitk.GetImageFromArray(img_array2[st:en, :, :].astype(np.int16))
        os.remove(se2output)
        sitk.WriteImage(out2, se2output)

        os.remove(se0output)
        os.remove(se1output)
        out0 = sitk.GetImageFromArray(img_array0[st:en, :, :].astype(np.int16))
        sitk.WriteImage(out0, se0output)
        out1 = sitk.GetImageFromArray(img_array1[st:en, :, :].astype(np.int16))
        sitk.WriteImage(out1, se1output)

        i = i + 1
        if i % 10 == 0:
            print('numbers:', i)

def resample_sitkimage(sitk_image_,new_size_):
    org_spacing = sitk_image_.GetSpacing()
    org_size = sitk_image_.GetSize()
    print(f'before resize, spacing: {org_spacing}, size: {org_size}')
    new_space = [(old_size*old_space/new_size) for old_size,old_space,new_size in zip(org_size,org_spacing,new_size_)]

    resampled_img=sitk.Resample(
        sitk_image_,
        new_size_,
        sitk.Transform(),
        sitk.sitkLinear,
        sitk_image_.GetOrigin(),
        new_space,
        sitk_image_.GetDirection(),
        0,
        sitk_image_.GetPixelID()
        )
    return resampled_img

def resample(path):
    input_nifti = sitk.ReadImage(path)# 读取nii.gz文件
    original_size = input_nifti.GetSize()  #获取原始图像尺寸
    if original_size[0]!=800 or original_size[1]!=800:
        print(path)
        z=min(original_size[0],original_size[1],original_size[2])
        target_size = [800, 800,z]# 定义目标尺寸为(原Z轴大小, 512, 512)
        resized_image = resample_sitkimage(input_nifti,target_size)
        sitk.WriteImage(resized_image, path)# 写入新的nii.gz文件
    if original_size[0]==800 and original_size[1]==800:
        img_array = sitk.GetArrayFromImage(input_nifti)
        out = sitk.GetImageFromArray(img_array.astype(np.int16))
        sitk.WriteImage(out, path)
    else:
        print("have some problem!")
        a=1
        # print("it's ok")

def set_direction(path):
    input_nifti = sitk.ReadImage(path)  # 读取nii.gz文件
    img_array = sitk.GetArrayFromImage(input_nifti)
    original_size=img_array.shape #z y x
    minn=min(original_size)
    # original_size = input_nifti.GetSize()  # 获取原始图像尺寸 x y z,对应于numpy中的z y x
    if original_size[2] == minn:
        img_array=img_array.transpose(2, 1, 0)
    elif original_size[1] == minn:
        img_array=img_array.transpose(1, 0, 2)
        print(path)

    zoom_factors = (1, 800 / original_size[1], 800 / original_size[2])
    resampled_image = zoom(img_array, zoom_factors, order=1)
    #resampled_image = np.flip(resampled_image, axis=1)  #处理external时候使用
    # resampled_image = resampled_image.transpose(0, 2, 1)
    # target_size = [minn, 800, 800]
    # resampled_image = resize(img_array, target_size, anti_aliasing=True)
    out = sitk.GetImageFromArray(resampled_image.astype(np.int16))
    sitk.WriteImage(out, path)

def process_breast():
    path = "/media/bit301/data/yml/data/xy/breast/set2"#/jfj/p1/dis/dml/fb/DICOM0/PA0
    path_list=[]
    for root, dirs, files in os.walk(path, topdown=False):
        if "0.nii.gz" in files:
            file_path = os.path.join(root, "0.nii.gz")
            path_list.append(file_path)

    for j in range(len(path_list)):
        # if "normal" not in path_list[j]:
        #     input_nifti1 = sitk.ReadImage(path_list[j].replace("0.nii.gz", "1.nii.gz"))  # 读取nii.gz文件
        # else:
        #     img_array0=np.zeros([original_size[0],original_size[1],original_size[2]])
        #     blank_label = sitk.GetImageFromArray(img_array0)
        #     sitk.WriteImage(blank_label, path_list[j].replace("0.nii.gz", "1.nii.gz"))
        set_direction(path_list[j])
        set_direction(path_list[j].replace("0.nii.gz","1.nii.gz"))
        if j % 10 == 0:
            print('numbers:', j)
    print("finished！")
    
def read_types():
    # "./subtype_data_20240413.xlsx"   +3+400+14
    #"./external_test_GE20240605_2.xlsx" +3+400+6
    path = "./external_test_GE20240605_2.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = str(sheet.cell(row=i, column=1).value)
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=6).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1

    # path = "./Clinical_and_Other_Features_20230420.xlsx"
    # wb = openpyxl.load_workbook(path)
    # sheet = wb[wb.sheetnames[0]]
    # for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
    #     ID = sheet.cell(row=i, column=1).value
    #     # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
    #     sub_data1 = sheet.cell(row=i, column=13).value
    #     if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
    #         type_dict[ID] = sub_data1
    return type_dict

#make the list of nnunet to 3DUxnet

def process_mask():
    types=read_types()
    # path = "/media/bit301/data/yml/data/xy/breast/set1"#/jfj/p1/dis/dml/fb/DICOM0/PA0
    path = "/media/bit301/data/yml/data/xy/breast/external"
    i=0
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            # if i<180:
            #     i=i+1
            #     if i % 10 == 0:
            #         print('numbers:', i)
            #     continue
            if "1.nii.gz" in path and "normal" not in path:
                    key=path.split('/')[-2]
                    typ=types[key]
                    # print(typ)
                    read = sitk.ReadImage(path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
                    img_array0 = sitk.GetArrayFromImage(read)
                    img_array0=np.where(img_array0>0,1,0)
                    img_array0 = img_array0*typ
                    out = sitk.GetImageFromArray(img_array0.astype(np.int16))
                    os.remove(path)
                    sitk.WriteImage(out, path)
                    i = i + 1
                    if i % 10 == 0:
                        print('numbers:', i)

#对open set 进行处理
def mv_file():
    path0 = "/media/bit301/data/yml/data/xy/breast/open_mask"
    path1 = "/media/bit301/data/yml/data/xy/breast/open"#/jfj/p1/dis/dml/fb/DICOM0/PA0
    path1_list=[]
    ii = 0
    for root, dirs, files in os.walk(path1, topdown=False):
        if "0.nii.gz" in files:
            path1_list.append(root)
            file=root.split("_")[-1]+".nii.gz"
            file_path = os.path.join(path0, file)
            if not os.path.exists(file_path):#文件不存在，跳过
                shutil.rmtree(root)
                continue
            target_path=os.path.join(root, "1.nii.gz")
            # shutil.move(file_path, root)
            shutil.copy2(file_path, target_path)#复制并重新命名
            ii = ii + 1
            if ii % 10 == 0:
                print('numbers:', ii)
            # if not os.path.isdir(catch):
            #     os.makedirs(catch)

def get_files_list():
    path = "/media/bit301/data/yml/data/xy/breast/set1"#
    files_list = "internal_test.txt"
    f1 = open(files_list, "w")  # 564
    path_list=[]
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path and "normal" not in path:
            # if "0.nii.gz" in path:
                path_list.append(path)
    path_list.sort()
    for j in range(len(path_list)):
        f1.writelines(path_list[j] + "\n")
    # ff.close()
    f1.close() #关

#按照类型划分训练-验证-测试数据
def split():
    types=read_types()
    value_counts = Counter(types.values())
    sorted_counts = sorted(value_counts.items(), key=lambda x: x[0], reverse=False)
    path = "/media/bit301/data/yml/data/xy/breast/set1"#
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path and "normal" not in path:
                path_list.append(path)
    random.seed(0)
    random.shuffle(path_list)
    cross=5
    k=0
    # train = "train.txt"
    # test = "test.txt"
    # f1 = open(train, "w")  # 564
    # f2 = open(test, "w")  # 188
    for i in range(cross):
        # if i==1:
        #     break
        train="train"+str(i)+".txt"
        test = "val" + str(i) + ".txt"
        f1 = open(train,"w")  # 564
        f2 = open(test, "w")#188
        for num, count in sorted_counts:
            # a = np.ceil(count / cross)
            # a = count // cross
            a=round(count / cross)
            # b=count % cross
            m = 0
            for j in range(len(path_list)):
                path=path_list[j]
                key = path.split('/')[-2]
                typ = types[key]
                if typ==num:
                    if m >= (cross - 1 - i) * a and m <(cross - i) * a:
                        f2.writelines(path + "\n")
                    else:
                        f1.writelines(path + "\n")
                    m=m+1
        f1.close()  # 关
        f2.close()

def check_split():
    types=read_types()
    value_counts = Counter(types.values())
    sorted_counts = sorted(value_counts.items(), key=lambda x: x[0], reverse=False)
    path = "/media/bit301/data/yml/data/xy/breast/set1"#
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path and "normal" not in path:
                path_list.append(path)
    cross=5
    val=[]
    for i in range(cross):
        test = "val" + str(i) + ".txt"
        sample_list = open(test).readlines()
        s = [x.strip() for x in sample_list]
        val.extend(s)

    for j in range(len(path_list)):
        path = path_list[j]
        if path not in val:
            print(path)

#按照类型划分训练-验证-测试数据
def patch():
    import copy
    types=read_types()
    value_counts = Counter(types.values())
    sorted_counts = sorted(value_counts.items(), key=lambda x: x[0], reverse=False)
    path = "/media/bit301/data/yml/data/xy/breast/set1"#
    path_list = []
    for root, dirs, files in os.walk(path, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "1.nii.gz" in path and "normal" not in path:
                path_list.append(path)
    random.seed(0)
    random.shuffle(path_list)
    patch = "patch_external.txt"
    f1 = open(patch, "w")  # 564
    ii=0
    for j in range(len(path_list)):
        path = path_list[j]
        label = sitk.ReadImage(path)
        label = sitk.GetArrayFromImage(label)
        label = label.astype(np.float32)
        label_size = label.shape
        left_mask = copy.deepcopy(label)
        right_mask = copy.deepcopy(label)
        left_mask[:, :, label_size[1] // 2:] = 0
        right_mask[:, :, :label_size[1] // 2] = 0
        if np.sum(left_mask) > 0 and np.sum(right_mask) > 0:  # 左右两侧均有病灶时，随机选择一侧
            if np.sum(left_mask) > 2 * np.sum(right_mask):
                label = left_mask
            elif 2 * np.sum(left_mask) < np.sum(right_mask):
                label = right_mask
            else:
                if random.random() > 0.5:
                    label = left_mask
                else:
                    label = right_mask

        index = np.nonzero(label)
        index = np.transpose(index)

        z_min = np.min(index[:, 0])
        z_max = np.max(index[:, 0])
        y_min = np.min(index[:, 1])
        y_max = np.max(index[:, 1])
        x_min = np.min(index[:, 2])
        x_max = np.max(index[:, 2])
        z=z_max-z_min
        y=y_max-y_min
        x=x_max-x_min
        # if x>128 or y>128:
        out=path.split("/")[-2]+" "+str(z)+" "+str(y)+" "+str(x)
        f1.writelines(out + "\n")
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    f1.close()

if __name__ == '__main__':
    # copy_and_paste()
    # display()
    # make_mask()

    # crop_by_mask()
    get_files_list()
    # json_to_list()
    # mv_file()
    # process_breast()
    # process_mask()
    # get_files_list()
    # split()
    # check_split()
    # patch()
    a=1
