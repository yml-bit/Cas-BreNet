import random
from sklearn.metrics import confusion_matrix
from sklearn.metrics import cohen_kappa_score
import openpyxl
import numpy as np

def read_types():
    path = "./internal_radiologists.xlsx"  # "./subtype_data_20240413.xlsx"
    # path = "./external_radiologists.xlsx"
    look_up_table_row_start = 2
    look_up_table_row_number = 200  # 根据实际行数调整
    type_dict_t = {}
    type_dict_l = {}
    type_dict_h = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = str(sheet.cell(row=i, column=1).value)#不加str，遇到纯数值时候会报错，
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=2).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict_t[ID] = sub_data1
        sub_data2 = sheet.cell(row=i, column=3).value
        if isinstance(sub_data2, (str, int, float)) and not np.isnan(sub_data2):
            type_dict_l[ID] = sub_data2
        sub_data3 = sheet.cell(row=i, column=4).value
        if isinstance(sub_data3, (str, int, float)) and not np.isnan(sub_data3):
            type_dict_h[ID] = sub_data3

    return type_dict_t,type_dict_l,type_dict_h
def calculate_metrics(confusion_matrix):
    # 2-TP/TN/FP/FN的计算
    weight=confusion_matrix.sum(axis=0)/confusion_matrix.sum()## 求出每列元素的和
    FN = confusion_matrix.sum(axis=0) - np.diag(confusion_matrix)
    FP = confusion_matrix.sum(axis=1) - np.diag(confusion_matrix)
    TP = np.diag(confusion_matrix)#所有对的 TP.sum=TP+TN
    TN = confusion_matrix.sum() - (FP + FN + TP)
    FP = FP.astype(float)
    FN = FN.astype(float)
    TP = TP.astype(float)
    TN = TN.astype(float)
    # print(TP)
    # print(TN)
    # print(FP)
    # print(FN)

    # 3-其他的性能参数的计算
    TPR = TP / (TP + FN)  # Sensitivity/ hit rate/ recall/ true positive rate 对于ground truth
    TNR = TN / (TN + FP)  # Specificity/ true negative rate  对于
    PPV = TP / (TP + FP)  # Precision/ positive predictive value  对于预测而言
    NPV = TN / (TN + FN)  # Negative predictive value
    FPR = FP / (FP + TN)  # Fall out/ false positive rate
    FNR = FN / (TP + FN)  # False negative rate
    FDR = FP / (TP + FP)  # False discovery rate
    sub_ACC = TP / (TP + FN)  # accuracy of each class
    acc=(TP+TN).sum()/(TP+TN+FP+FN).sum()
    average_acc=TP.sum() / (TP.sum() + FN.sum())
    F1_Score=2*TPR*PPV/(PPV+TPR)
    Macro_F1=F1_Score.mean()
    return TPR.mean(), TNR.mean(), average_acc,Macro_F1

if __name__ == '__main__':
    types=read_types()#注意修改
    correct = 0
    total = 0
    file_path="./internal_test.txt"#internal_test external_test
    # file = "./internal_radiologist_l.txt"
    file = "./internal_radiologist_l.txt"
    f_test = open(file, "w")
    y_true = []
    y_pred = []
    predictions = {}
    with open(file_path, 'r') as file:
        for line in file:
            id=line.split("/")[-2]
            # id = int(id)
            pred = int(types[1][id])#l=types[1][id] h=types[2][id]
            if 0.5 <= pred <= 2.5:
                pred = 1
            elif 2.5 < pred <= 4.5:
                pred = 2
            else:
                pred = 3
                
            true = int(types[0][id])#l
            if 0.5 <= true <= 2.5:
                true = 1
            elif 2.5 < true <= 4.5:
                true = 2
            else:
                true = 3

            line = str(id) + "    " + str(pred) + "    " + str(true)
            # predicted = Counter(predicted_logits).most_common()[0][0]
            # line = str(batch2["ID"]) + "    " + str(values_list[0]) + "    " + str(predicted)
            f_test.writelines(str(line) + "\n")
            y_true.append(true)
            y_pred.append(pred)
        confm = confusion_matrix(y_true, y_pred)
        ckap = cohen_kappa_score(y_true, y_pred)
        Sensitivity, specificity, accuracy, f1_macro = calculate_metrics(confm)
        # f1_macro = f1_score(y_true, y_pred, average='macro')
        print("Current  Avg. confm: \n{} ".format(confm))
        print("Current  Avg. ckap: {} ".format(ckap))
        print("Current  Avg. acc: {} ".format(accuracy))
        print("Current  Avg. Sensitivity: {} ".format(Sensitivity))
        print("Current  Avg. specificity: {} ".format(specificity))
        print("Current  Avg. f1_score: {} ".format(f1_macro))
        f_test.close()