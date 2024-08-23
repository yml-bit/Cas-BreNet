import torch
torch.multiprocessing.set_sharing_strategy('file_system')##训练python脚本中import torch后，加上下面这句。
from networks.Mutitask.network_backbone import unet
from networks.SegResNet.network_backbone import SegResNet
from networks.Unet_3D.network_backbone import BasicUNet
from networks.Mednext.MedNextV1 import MedNeXt
from networks.Mednext.CNN3D import CNN3D
from networks.Mednext.MedCNN import MedCNN
from networks.Mednext.MSCNN3D import MSCNN3D
from networks.nSwinunetr.nSwinunetr import nSwinUNETR
from monai.data import CacheDataset, DataLoader, decollate_batch
from monai.metrics import DiceMetric
from datasets.load_datasets_transforms import dataset, Generator1
from torchvision import transforms
import openpyxl
from monai.utils import set_determinism
from sklearn.metrics import confusion_matrix
from sklearn.metrics import cohen_kappa_score


import os
import argparse
import yaml
import random
import numpy as np
import SimpleITK as sitk

def config():
    parser = argparse.ArgumentParser(description='3D UX-Net hyperparameters for medical image segmentation')
    ## Input data hyperparameters
    # parser.add_argument('--root', type=str, default='', required=True, help='Root folder of all your images and labels')
    parser.add_argument('--dataset', type=str, default='301', help='Datasets: {feta, flare, amos}, Fyi: You can add your dataset here')
    parser.add_argument('--train_list', type=str, default='./data/train3.txt',help='Root folder')
    parser.add_argument('--val_list', type=str, default='./data/val3.txt', help='Root folder')#因为要测试评分，所以使用validation代码
    parser.add_argument('--test_list', type=str, default='./data/test.txt',help='Root folder')
    parser.add_argument('--list_dir', type=str, default='./data/',help='Root folder')
    parser.add_argument('--seed', type=int,default=1234, help='random seed')
    parser.add_argument('--output', type=str, default='./output/MedNeXtm', help='Output folder for both tensorboard and the best model')
    parser.add_argument('--image_save', type=str, default='./output/MedNeXtm/img/', help='images')
    ## Input model & training hyperparameters
    parser.add_argument('--network', type=str, default='MedNeXt',help='Network models: {TransBTS, nnFormer, UNETR, SwinUNETR, 3DUXNET}')
    parser.add_argument('--supervision', type=bool,default=False, help='supervision')#True False
    parser.add_argument('--trained_weights', default='./output/MedNeXt/0604', help='Path of pretrained/fine-tuned weights')
    parser.add_argument('--patch', type=int, default=(48,96,96), help='Batch size for subject input')
    parser.add_argument('--num_classes', type=int, default=6, help='Number of classes')
    parser.add_argument('--mode', type=str, default='val', help='Training or testing mode')
    parser.add_argument('--batch_size', type=int, default=1, help='Batch size for subject input')
    parser.add_argument('--crop_sample', type=int, default=1, help='Number of cropped sub-volumes for each subject')
    parser.add_argument('--sw_batch_size', type=int, default=4, help='Sliding window batch size for inference')
    parser.add_argument('--overlap', type=float, default=0.5, help='Sub-volume overlapped percentage')

    ## Efficiency hyperparameters
    parser.add_argument('--gpu', type=str, default='0', help='your GPU number')
    parser.add_argument('--cache_rate', type=float, default=1.0, help='Cache rate to cache your dataset into GPUs')#0.1
    parser.add_argument('--num_workers', type=int, default=1, help='Number of workers')

    args = parser.parse_args()
    return args

def read_types():
    path = "./data/subtype_data_20240413.xlsx"  # "./subtype_data_20240413.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=14).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1

    path = "./data/Clinical_and_Other_Features_20230420.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=13).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1
    return type_dict

def get_config(config):
    with open(config, 'r') as stream:
        return yaml.load(stream)

def seed_everything(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    #torch.backends.cudnn.benchmark = True / False

def compute_metrics(a, b):
    """
    计算敏感性（Recall）、特异性（Specificity）和准确率（Accuracy）。

    参数:
    a: 真实标签列表，元素为0（负类）或1（正类）
    b: 预测结果列表，元素为0（负类）或1（正类）

    返回:
    recall: 敏感性
    specificity: 特异性
    accuracy: 准确率
    """
    # 确保a和b长度相同
    assert len(a) == len(b), "真实标签和预测结果的长度不一致"

    # 初始化混淆矩阵计数器
    TP = TN = FP = FN = 0

    # 遍历预测结果和真实标签，统计混淆矩阵各部分的数量
    for pred, true in zip(b, a):
        if pred == 1 and true == 1:
            TP += 1
        elif pred == 0 and true == 0:
            TN += 1
        elif pred == 1 and true == 0:
            FP += 1
        else:  # pred == 0 and true == 1
            FN += 1

    # 防止除以零错误
    def safe_divide(numerator, denominator):
        return numerator / denominator if denominator != 0 else 0.0

    # 计算敏感性（召回率）
    recall = safe_divide(TP, TP + FN)

    # 计算特异性
    specificity = safe_divide(TN, TN + FP)

    # 计算准确率
    accuracy = safe_divide(TP + TN, TP + TN + FP + FN)

    return recall, specificity, accuracy

if __name__ == '__main__':
    # seed_everything(seed=42)
    set_determinism(seed=1234)
    types=read_types()
    args=config() #the first running should excute this code
    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

    # aa = np.random.rand(512,512,370)
    # aa = aa[np.newaxis, np.newaxis, :, :, :]
    # ab = torch.tensor(aa)  # .cuda()
    # ac = MetaTensor(ab).cuda()

    db_val = dataset(list_dir=args.val_list, split="val", num_classes=args.num_classes,
                               transform=transforms.Compose([Generator1(output_size=args.patch, mode = 'val')]))
    print("The length of val set is: {}".format(len(db_val)))
    def worker_init_fn(worker_id):
        random.seed(args.seed + worker_id)
    val_loader = DataLoader(db_val, batch_size=1, shuffle=True, num_workers=6, pin_memory=True,
                             worker_init_fn=worker_init_fn)

    ## Load Networks
    ## Load Networks
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    patch=np.array(args.patch,dtype=int) #(96,96,48)
    patch = np.array(args.patch, dtype=int)  # (96,96,48)
    if args.network =='unet':##(1,1,96,96,96)
        model = unet().to(device)
    elif args.network == 'SegResNet':
        model = SegResNet().to(device)
    elif args.network == 'BasicUNet':
        model = BasicUNet().to(device)
    elif args.network == "MedNeXt":  ##(160,160,96)
        model = MedNeXt(
       in_channels = 1,
        n_channels = 32,
        n_classes = args.num_classes,
        exp_r=[2,3,4,4,4,4,4,3,2],
        kernel_size=5,
        deep_supervision=args.supervision,
        do_res=True,
        do_res_up_down = True,
        block_counts = [3,4,4,4,4,4,4,4,3],
        checkpoint_style = 'outside_block'
       #  in_channels=1,
       #  n_channels=32,
       #  n_classes=args.num_classes,
       #  exp_r=[3, 4, 8, 8, 8, 8, 8, 4, 3],
       #  kernel_size=3,
       #  deep_supervision=args.supervision,
       #  do_res=True,
       #  do_res_up_down=True,
       #  block_counts=[3, 4, 8, 8, 8, 8, 8, 4, 3],
       #  checkpoint_style='outside_block'
        ).to(device)
    elif args.network == "nSwinUNETR":  # (160,160,96)
        model = nSwinUNETR(
            img_size=args.patch,
            in_channels=1,
            out_channels=args.num_classes,
            feature_size=48,
            drop_rate=0.0,
            attn_drop_rate=0.0,
            dropout_path_rate=0.0,
            use_checkpoint=True,
            deep_supervision=args.supervision,
        ).to(device)
    elif args.network == "MSCNN3D":  ##(160,160,96)
        model = MedCNN(args.num_classes).to(device)
    elif args.network == "CNN3D":  ##(160,160,96)
        model = CNN3D(
            in_channels=2,
            out_channels=64,
            num_classes=args.num_classes
        ).to(device)
    print('Chosen Network Architecture: {}'.format(args.network))
    model.load_state_dict(torch.load(args.trained_weights+"a1/best.pth"))
    model.eval()
    model2=model

    model2.load_state_dict(torch.load(args.trained_weights+"a2/best.pth"))
    model2.eval()

    model3=model
    model3.load_state_dict(torch.load(args.trained_weights+"a3/best.pth"))
    model3.eval()
    # dice_metric = DiceMetric(include_background=True, reduction="mean", get_not_nans=False)
    dice_metric = DiceMetric(include_background=False, reduction="none", get_not_nans=False)  # 去除背景项目
    dice_vals = list()
    correct = 0
    total = 0
    file="./data/record.txt"
    f_test=open(file,"w")
    y_true=[]
    y_pred=[]
    with torch.no_grad():
        for step, batch in enumerate(val_loader):
            torch.cuda.empty_cache()
            # type = types[batch["ID"]].long()
            values_list = [types[key] for key in batch["ID"] if key in types]
            catageries = torch.Tensor(values_list).long().to(device)
            patch = batch["image"]
            predict_type1 = model(patch.to(device))
            predict_type2 = model2(patch.to(device))
            predict_type3 = model3(patch.to(device))
            size=patch.shape
            aa=size(-1)//32-3
            if aa==1:
                predict_type=3*predict_type1+predict_type2+predict_type3
            elif aa==2:
                predict_type=2*predict_type1+3*predict_type2+predict_type3
            else:
                predict_type = 1 * predict_type1 + 2 * predict_type2 + 3*predict_type3
            _, predicted = torch.max(predict_type.data, 1)
            total += catageries.size(0)
            correct += (predicted == catageries).sum().item()
            line=str(batch["ID"])+"    "+str(values_list[0])+"    "+str(predicted.item())
            f_test.writelines(str(line)+"\n")
            y_true.append(values_list[0])
            y_pred.append(predicted.item())
        confm=confusion_matrix(y_true, y_pred)
        ckap= cohen_kappa_score(y_true, y_pred)
        Sensitivity, specificity, accuracy = compute_metrics(y_true, y_pred)
        precision = correct / total
        # Sensitivity = correct / total
        f1_score = 2 * (precision * Sensitivity) / (precision + Sensitivity)
        print("Current  Avg. confm: {} ".format(confm))
        print("Current  Avg. ckap: {} ".format(ckap))
        print("Current  Avg. acc: {} ".format(precision))
        print("Current  Avg. Sensitivity: {} ".format(Sensitivity))
        print("Current  Avg. specificity: {} ".format(specificity))
        print("Current  Avg. f1_score: {} ".format(f1_score))
        f_test.close()