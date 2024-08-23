import torch
torch.multiprocessing.set_sharing_strategy('file_system')##训练python脚本中import torch后，加上下面这句。
import torch.nn as nn
import numpy as np
import os
from tqdm import tqdm
import argparse
import datetime
from monai.utils import set_determinism

from networks.Mednext.MedNextV1_MT import MedNeXt_MT
from monai.transforms import AsDiscrete
from monai.metrics import DiceMetric
from monai.losses import DiceCELoss,DiceFocalLoss
from monai.inferers import sliding_window_inference
from monai.data import CacheDataset, DataLoader, decollate_batch,PersistentDataset
from torch.utils.tensorboard import SummaryWriter
from openpyxl import load_workbook
from datasets.load_datasets_transforms_M import data_transforms,dataset, RandomGenerator,Generator,patch_extraction,patch_reconstruction,roi_extraction
from torchvision import transforms
import random
from torchmetrics import Accuracy
from collections import Counter
import warnings
import openpyxl
# 忽略所有警告
warnings.filterwarnings("ignore")

def config():
    parser = argparse.ArgumentParser(description='hyperparameters for medical image segmentation and classification')
    parser.add_argument('--dataset', type=str, default='301', help='Datasets: {feta, flare, amos}, Fyi: You can add your dataset here')
    parser.add_argument('--train_list', type=str, default='./data/train4.txt',help='Root folder')
    parser.add_argument('--val_list', type=str, default='./data/val4.txt', help='Root folder')
    parser.add_argument('--list_dir', type=str, default='./data/',help='Root folder')
    parser.add_argument('--seed', type=int,default=1234, help='random seed')
    #unet SegResNet BasicUNet
    parser.add_argument('--output', type=str, default='./output/MedNeXt_M/0528e/', help='Output folder for both tensorboard and the best model')
    parser.add_argument('--image_save', type=str, default='./output/MedNeXt_M/0528e/', help='images')
    parser.add_argument('--network', type=str, default='MedNeXt', help='Network models: {TransBTS, nnFormer, UNETR, SwinUNETR, 3DUXNET}')
    parser.add_argument('--supervision', type=bool,default=False, help='supervision')#True False
    parser.add_argument('--mode', type=str, default='train', help='Training or testing mode')
    parser.add_argument('--pretrain', default="True", help='Have pretrained weights or not')
    parser.add_argument('--pretrained_weights', default='', help='Path of pretrained weights')
    #0405+(64,96,96)+0.59375;0406+(64,96,96)+0.59375；0423+(64,96,96)+0.59375;
    #0424+(64,128,128)+0.5;追加15例数据
    #0425+(64,96,96)+0.405；使用开源数据训练
    #BasicUNet+0425+(64,96,96)+0.3125；使用开源数据训练
    parser.add_argument('--patch', type=int, default=(48,160,160), help='Batch size for subject input')#(128,160,160) (80,128,128)
    parser.add_argument('--num_classes', type=int, default=6, help='Number of classes')
    parser.add_argument('--batch_size', type=int, default=4, help='Batch size for subject input')
    parser.add_argument('--crop_sample', type=int, default=1, help='Number of cropped sub-volumes for each subject')
    parser.add_argument('--lr', type=float, default=3e-4, help='Learning rate for training')
    parser.add_argument('--optim', type=str, default='AdamW', help='Optimizer types: Adam / AdamW')
    parser.add_argument('--max_iter', type=int, default=8000, help='Maximum iteration steps for training')
    parser.add_argument('--eval_step', type=int, default=500, help='Per steps to perform validation')#4000 1e8

    ## Efficiency hyperparameters
    parser.add_argument('--gpu', type=str, default='0', help='your GPU number')
    parser.add_argument('--cache_rate', type=float, default=0.5, help='Cache rate to cache your dataset into GPUs')#0.1  缓存数据占总数的百分比!!
    parser.add_argument('--num_workers', type=int, default=6, help='Number of workers')

    args = parser.parse_args()
    # with open('./Yaml/3DX-Unet1.yaml', 'w') as f:
    #     yaml.dump(args.__dict__, f, indent=2)
    return args

def read_types():
    path = "./data/subtype_data_20240413.xlsx"  # "./subtype_data_20240413.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=14).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1

    # path = "./data/Clinical_and_Other_Features_20230420.xlsx"
    # wb = openpyxl.load_workbook(path)
    # sheet = wb[wb.sheetnames[0]]
    # for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
    #     ID = sheet.cell(row=i, column=1).value
    #     # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
    #     sub_data1 = sheet.cell(row=i, column=13).value
    #     if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
    #         type_dict[ID] = sub_data1
    return type_dict

def seed_everything(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    #torch.backends.cudnn.benchmark = True / False

if __name__ == '__main__':
    # seed_everything(seed=42)
    set_determinism(seed=1234)
    types=read_types()
    args=config() #the first running should excute this code
    # device="cpu"
    # os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu
    # print('Used GPU: {}'.format(args.gpu))
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print('Used GPU: {}'.format(device))
    train_transforms, val_transforms = data_transforms(args)
    
    # db_train = dataset(list_dir=args.list_dir, split="train", num_classes=args.num_classes,
    #                            transform=train_transforms)#0225
    db_train = dataset(list_dir=args.train_list, split="train", num_classes=args.num_classes,
                       transform=transforms.Compose([RandomGenerator(output_size=args.patch, mode='train')]))#0226
    print("The length of train set is: {}".format(len(db_train)))
    db_val = dataset(list_dir=args.val_list, split="val", num_classes=args.num_classes,
                               transform=transforms.Compose([Generator(output_size=args.patch, mode = 'val')]))
    print("The length of val set is: {}".format(len(db_val)))

    def worker_init_fn(worker_id):
        random.seed(args.seed + worker_id)

    train_loader = DataLoader(db_train, batch_size=args.batch_size, shuffle=True, num_workers=6, pin_memory=True,
                             worker_init_fn=worker_init_fn)
    val_loader = DataLoader(db_val, batch_size=1, shuffle=True, num_workers=6, pin_memory=True,
                             worker_init_fn=worker_init_fn)

    model = MedNeXt_MT(
        in_channels=1,
        n_channels=32,
        n_classes=args.num_classes,
        exp_r=[2, 3, 4, 4, 4, 4, 4, 3, 2],
        kernel_size=5,
        deep_supervision=args.supervision,
        do_res=True,
        do_res_up_down=True,
        block_counts=[3, 4, 4, 4, 4, 4, 4, 4, 3],
        checkpoint_style='outside_block'
        # in_channels=1,
        # n_channels=32,
        # n_classes=out_classes,
        # exp_r=[3, 4, 8, 8, 8, 8, 8, 4, 3],
        # kernel_size=3,
        # deep_supervision=args.supervision,
        # do_res=True,
        # do_res_up_down=True,
        # block_counts=[3, 4, 8, 8, 8, 8, 8, 4, 3],
        # checkpoint_style='outside_block'
    ).to(device)
    print('Chosen Network Architecture: {}'.format(args.network))
    
    if args.pretrain == 'False':# True False
        print('Pretrained weight is found! Start to load weight from')
        modd = "02_22_12.pth"
        model.load_state_dict(torch.load(os.path.join(args.output, modd)))
    loss1 = DiceCELoss(to_onehot_y=True, softmax=True)  # class_weight
    loss2 = nn.CrossEntropyLoss()
    # loss_function = DiceFocalLoss(to_onehot_y=True, softmax=True)
    print('Loss for training: {}'.format('DiceCELoss'))
    if args.optim == 'AdamW':
        optimizer = torch.optim.AdamW(model.parameters(), lr=args.lr)
    elif args.optim == 'Adam':
        optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)
    print('Optimizer for training: {}, learning rate: {}'.format(args.optim, args.lr))
    # scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, 'min', factor=0.9, patience=1000)

    root_dir = os.path.join(args.output)
    if os.path.exists(root_dir) == False:
        os.makedirs(root_dir)

    t_dir = os.path.join(root_dir, 'tensorboard')
    if os.path.exists(t_dir) == False:
        os.makedirs(t_dir)
    writer = SummaryWriter(log_dir=t_dir)

    def validation(epoch_iterator_val):
        model.eval()
        pred=[]
        gd=[]
        with torch.no_grad():
            for step, batch in enumerate(epoch_iterator_val):
                torch.cuda.empty_cache()
                values_list = [types[key] for key in batch["ID"] if key in types]
                catageries = torch.Tensor(values_list).long()
                patchs=batch["image"]
                for patch in patchs:
                    logit_map,predict_type = model(patch.to(device))
                    _, predicted_classes = torch.max(predict_type, dim=1)
                    pred.append(predicted_classes.cpu().detach())
                    gd.append(catageries)
        print("number of diagnosis:"+str(len(gd)))
        mean_dis_val = accuracy_metric(torch.tensor(pred),torch.tensor(gd))
        # mean_dis_val = np.nanmean(dis_vals)
        writer.add_scalar('Validation Cross Loss', mean_dis_val, global_step)
        return mean_dis_val

    def train(global_step, train_loader, dice_val_best,dis_val_best, global_step_best):
        model.train()
        epoch_loss = 0
        step = 0
        epoch_iterator = tqdm(train_loader, desc="Training (X / X Steps) (loss=X.X)", dynamic_ncols=True)
        for step, batch in enumerate(epoch_iterator):
            step += 1
            values_list = [types[key] for key in batch["ID"] if key in types]
            catageries=torch.Tensor(values_list).long()
            catageries=catageries*batch["flag"]
            # image_batch, label_batch = batch['image'], batch['label']
            x, y,z = (batch["image"].to(device), batch["label"].to(device),catageries.to(device))
            logit_map,predict_type = model(x)
            loss=loss1(logit_map, y)+loss2(predict_type,z)
            loss.backward()
            epoch_loss += loss.item()
            optimizer.step()
            optimizer.zero_grad()
            epoch_iterator.set_description("Training (%d / %d Steps) (loss=%2.5f)" % (global_step, max_iterations, loss)
            )
            if (
                    global_step % eval_num == 0 and global_step != 0
            ) or global_step == max_iterations:
                epoch_iterator_val = tqdm(
                    val_loader, desc="Validate (X / X Steps) (dice=X.X)", dynamic_ncols=True
                )
                dis_val = validation(epoch_iterator_val)
                epoch_loss /= step
                epoch_loss_values.append(epoch_loss)
                metric_values.append(dis_val)
                if dis_val > dis_val_best:
                    dis_val_best = dis_val
                    global_step_best = global_step
                    modd = str(datetime.datetime.now().strftime('%m_%d_%H')) + ".pth"
                    torch.save(model.state_dict(), os.path.join(root_dir, modd))
                    print(
                        "Model Was Saved ! Current Best Avg. acc: {} Current Avg. acc: {}".format(
                            dis_val_best, dis_val))
                    # scheduler.step(dice_val)
                else:
                    print("Model Was Not Saved ! Current Best Avg. acc: {} Current Avg. acc: {}".format(
                            dis_val_best, dis_val))
                    # scheduler.step(dice_val)
            writer.add_scalar('Training Segmentation Loss', loss.data, global_step)
            global_step += 1
            del loss
        return global_step, dice_val_best,dis_val_best, global_step_best

    max_iterations = args.max_iter
    print('Maximum Iterations for training: {}'.format(str(args.max_iter)))
    eval_num = args.eval_step
    post_label = AsDiscrete(to_onehot=args.num_classes)
    post_pred = AsDiscrete(argmax=True, threshold=0.5,to_onehot=args.num_classes)
    dice_metric = DiceMetric(include_background=False, reduction="none", get_not_nans=False)  # 去除背景项目
    accuracy_metric = Accuracy(task="multiclass",num_classes=12)
    dice_val_best=0.0
    dis_val_best = 0.0
    global_step_best = 0
    epoch_loss_values = []
    metric_values = []
    global_step=0
    while global_step < max_iterations:
        global_step, dice_val_best,dis_val_best, global_step_best = train(
            global_step, train_loader, dice_val_best,dis_val_best, global_step_best
        )