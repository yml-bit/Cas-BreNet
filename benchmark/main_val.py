import torch
torch.multiprocessing.set_sharing_strategy('file_system')##训练python脚本中import torch后，加上下面这句。
from networks.Mutitask.network_backbone import unet
from networks.SegResNet.network_backbone import SegResNet
from networks.Unet_3D.network_backbone import BasicUNet
from networks.Mednext.MedNextV1 import MedNeXt
from networks.Mednext.CNN3D import CNN3D
from networks.Mednext.MedCNN import MedCNN
from networks.Mednext.MSCNN3D import MSCNN3D
from networks.nSwinunetr.nSwinunetr import nSwinUNETR
from monai.data import CacheDataset, DataLoader, decollate_batch
from monai.metrics import DiceMetric
from datasets.load_datasets_transforms import data_transforms,dataset, RandomGenerator,Generator,patch_extraction,patch_reconstruction,roi_extraction
from torchvision import transforms
import openpyxl
from monai.utils import set_determinism
from sklearn.metrics import confusion_matrix
from sklearn.metrics import cohen_kappa_score
from collections import Counter
from sklearn.metrics import f1_score


import os
import argparse
import yaml
import random
import numpy as np
import SimpleITK as sitk

def config():
    parser = argparse.ArgumentParser(description='3D UX-Net hyperparameters for medical image segmentation')
    ## Input data hyperparameters
    # parser.add_argument('--root', type=str, default='', required=True, help='Root folder of all your images and labels')
    parser.add_argument('--dataset', type=str, default='301', help='Datasets: {feta, flare, amos}, Fyi: You can add your dataset here')
    parser.add_argument('--train_list', type=str, default='./data/train1.txt',help='Root folder')
    #val0   external_list
    #CNN3D+(32,80,64)  MSCNN3D+(32,160,160)
    parser.add_argument('--val_list', type=str, default='./data/external_list.txt', help='Root folder')#因为要测试评分，所以使用validation代码
    parser.add_argument('--test_list', type=str, default='./data/test.txt',help='Root folder')
    parser.add_argument('--list_dir', type=str, default='./data/',help='Root folder')
    parser.add_argument('--seed', type=int,default=1234, help='random seed')
    parser.add_argument('--output', type=str, default='./output/CNN3D', help='Output folder for both tensorboard and the best model')
    parser.add_argument('--image_save', type=str, default='./output/MedNeXt/img/', help='images')
    ## Input model & training hyperparameters
    parser.add_argument('--network', type=str, default='CNN3D',help='Network models: {TransBTS, nnFormer, UNETR, SwinUNETR, 3DUXNET}')
    parser.add_argument('--supervision', type=bool,default=False, help='supervision')#True False
    parser.add_argument('--trained_weights', default='./output/CNN3D/0620e1/best.pth', help='Path of pretrained/fine-tuned weights')
    parser.add_argument('--patch', type=int, default=(32,80,64), help='Batch size for subject input')
    parser.add_argument('--num_classes', type=int, default=4, help='Number of classes')
    parser.add_argument('--mode', type=str, default='val', help='Training or testing mode')
    parser.add_argument('--batch_size', type=int, default=1, help='Batch size for subject input')
    parser.add_argument('--crop_sample', type=int, default=1, help='Number of cropped sub-volumes for each subject')
    parser.add_argument('--sw_batch_size', type=int, default=4, help='Sliding window batch size for inference')
    parser.add_argument('--overlap', type=float, default=0.5, help='Sub-volume overlapped percentage')

    ## Efficiency hyperparameters
    parser.add_argument('--gpu', type=str, default='0', help='your GPU number')
    parser.add_argument('--cache_rate', type=float, default=1.0, help='Cache rate to cache your dataset into GPUs')#0.1
    parser.add_argument('--num_workers', type=int, default=1, help='Number of workers')

    args = parser.parse_args()
    return args

def read_types1():
    path = "./data/subtype_data_20240413.xlsx"  # "./subtype_data_20240413.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=14).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1

    path = "./data/Clinical_and_Other_Features_20230420.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=13).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1
    return type_dict

def read_types():
    # "./subtype_data_20240413.xlsx"   +3+400+14
    #"./external_test_GE20240605_2.xlsx" +3+400+6
    path = "./data/external_test_GE20240605_2.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = str(sheet.cell(row=i, column=1).value)
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=6).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1
    return type_dict

def get_config(config):
    with open(config, 'r') as stream:
        return yaml.load(stream)

def seed_everything(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    #torch.backends.cudnn.benchmark = True / False

def calculate_metrics(confusion_matrix):
    # 2-TP/TN/FP/FN的计算
    weight=confusion_matrix.sum(axis=0)/confusion_matrix.sum()## 求出每列元素的和
    FN = confusion_matrix.sum(axis=0) - np.diag(confusion_matrix)
    FP = confusion_matrix.sum(axis=1) - np.diag(confusion_matrix)
    TP = np.diag(confusion_matrix)#所有对的 TP.sum=TP+TN
    TN = confusion_matrix.sum() - (FP + FN + TP)
    FP = FP.astype(float)
    FN = FN.astype(float)
    TP = TP.astype(float)
    TN = TN.astype(float)
    # print(TP)
    # print(TN)
    # print(FP)
    # print(FN)

    # 3-其他的性能参数的计算
    TPR = TP / (TP + FN)  # Sensitivity/ hit rate/ recall/ true positive rate 对于ground truth
    TNR = TN / (TN + FP)  # Specificity/ true negative rate  对于
    PPV = TP / (TP + FP)  # Precision/ positive predictive value  对于预测而言
    NPV = TN / (TN + FN)  # Negative predictive value
    FPR = FP / (FP + TN)  # Fall out/ false positive rate
    FNR = FN / (TP + FN)  # False negative rate
    FDR = FP / (TP + FP)  # False discovery rate
    sub_ACC = TP / (TP + FN)  # accuracy of each class
    acc=(TP+TN).sum()/(TP+TN+FP+FN).sum()
    average_acc=TP.sum() / (TP.sum() + FN.sum())
    F1_Score=2*TPR*PPV/(PPV+TPR)
    Macro_F1=F1_Score.mean()
    return TPR.mean(), TNR.mean(), average_acc,Macro_F1

if __name__ == '__main__':
    # seed_everything(seed=42)
    set_determinism(seed=1234)
    types=read_types()
    args=config() #the first running should excute this code
    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

    # aa = np.random.rand(512,512,370)
    # aa = aa[np.newaxis, np.newaxis, :, :, :]
    # ab = torch.tensor(aa)  # .cuda()
    # ac = MetaTensor(ab).cuda()

    db_val = dataset(list_dir=args.val_list, split="val", num_classes=args.num_classes,
                               transform=transforms.Compose([Generator(output_size=args.patch, mode = 'val')]))
    print("The length of val set is: {}".format(len(db_val)))
    def worker_init_fn(worker_id):
        random.seed(args.seed + worker_id)
    val_loader = DataLoader(db_val, batch_size=1, shuffle=True, num_workers=1, pin_memory=True,
                             worker_init_fn=worker_init_fn)

    ## Load Networks
    ## Load Networks
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    patch=np.array(args.patch,dtype=int) #(96,96,48)
    patch = np.array(args.patch, dtype=int)  # (96,96,48)
    if args.network =='unet':##(1,1,96,96,96)
        model = unet().to(device)
    elif args.network == 'SegResNet':
        model = SegResNet().to(device)
    elif args.network == 'BasicUNet':
        model = BasicUNet().to(device)
    elif args.network == "MedNeXt":  ##(160,160,96)
        model = MedNeXt(
       in_channels = 3,
        n_channels = 32,
        n_classes = args.num_classes,
        exp_r=[2,3,4,4,4,4,4,3,2],
        kernel_size=5,
        deep_supervision=args.supervision,
        do_res=True,
        do_res_up_down = True,
        block_counts = [3,4,4,4,4,4,4,4,3],
        checkpoint_style = 'outside_block'
       #  in_channels=1,
       #  n_channels=32,
       #  n_classes=args.num_classes,
       #  exp_r=[3, 4, 8, 8, 8, 8, 8, 4, 3],
       #  kernel_size=3,
       #  deep_supervision=args.supervision,
       #  do_res=True,
       #  do_res_up_down=True,
       #  block_counts=[3, 4, 8, 8, 8, 8, 8, 4, 3],
       #  checkpoint_style='outside_block'
        ).to(device)
    elif args.network == "nSwinUNETR":  # (160,160,96)
        model = nSwinUNETR(
            img_size=args.patch,
            in_channels=1,
            out_channels=args.num_classes,
            feature_size=48,
            drop_rate=0.0,
            attn_drop_rate=0.0,
            dropout_path_rate=0.0,
            use_checkpoint=True,
            deep_supervision=args.supervision,
        ).to(device)
    elif args.network == "MSCNN3D":  ##(160,160,96)
        model = MSCNN3D(args.num_classes).to(device)
    elif args.network == "CNN3D":
        model = CNN3D(
            in_channels=3,
            out_channels=64,
            num_classes=args.num_classes
        ).to(device)
    print('Chosen Network Architecture: {}'.format(args.network))
    model.load_state_dict(torch.load(args.trained_weights))
    model.eval()
    # dice_metric = DiceMetric(include_background=True, reduction="mean", get_not_nans=False)
    dice_metric = DiceMetric(include_background=False, reduction="none", get_not_nans=False)  # 去除背景项目
    dice_vals = list()
    correct = 0
    total = 0
    file="./output/CNN3D/external_e.txt"
    f_test=open(file,"w")
    y_true=[]
    y_pred=[]
    with torch.no_grad():
        for step, batch in enumerate(val_loader):
            torch.cuda.empty_cache()
            t = batch[0]["label"]
            catageries = torch.amax(t, dim=(1, 2, 3, 4))
            catageries[(catageries >= 0.5) & (catageries <= 2.5)] = 1
            catageries[(catageries >= 2.5) & (catageries <= 4.5)] = 2
            catageries[(catageries >= 4.5)] = 3
            catageries[(catageries <1)] = 3
            # catageries = torch.Tensor(catageries).long().to(device)
            patch = batch[0]["image"]
            # for patch in patchs:
            predict_type = model(patch.to(device))
            _, predicted = torch.max(predict_type.data, 1)
            out=predicted.item()
            if out<1:
                out=3
            line = str(batch[1]["ID"])  + "    " + str(out)+ "    " + str(int(catageries.item()))
            # predicted = Counter(predicted_logits).most_common()[0][0]
            # line = str(batch2["ID"]) + "    " + str(values_list[0]) + "    " + str(predicted)
            f_test.writelines(str(line) + "\n")
            y_true.append(catageries)
            y_pred.append(out)
        confm=confusion_matrix(y_true, y_pred)
        ckap= cohen_kappa_score(y_true, y_pred)
        Sensitivity, specificity, accuracy ,f1_macro= calculate_metrics(confm)
        # f1_macro = f1_score(y_true, y_pred, average='macro')
        print("Current  Avg. confm: \n{} ".format(confm))
        print("Current  Avg. ckap: {} ".format(ckap))
        print("Current  Avg. acc: {} ".format(accuracy))
        print("Current  Avg. Sensitivity: {} ".format(Sensitivity))
        print("Current  Avg. specificity: {} ".format(specificity))
        print("Current  Avg. f1_score: {} ".format(f1_macro))
        f_test.close()
