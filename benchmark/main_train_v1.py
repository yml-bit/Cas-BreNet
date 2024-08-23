import torch
torch.multiprocessing.set_sharing_strategy('file_system')##训练python脚本中import torch后，加上下面这句。
import torch.nn as nn
import numpy as np
import os
from tqdm import tqdm
import argparse
import datetime
from monai.utils import set_determinism

from networks.Mutitask.network_backbone import unet
from networks.SegResNet.network_backbone import SegResNet
from networks.Unet_3D.network_backbone import BasicUNet
from networks.Mednext.MedNextV1 import MedNeXt#(48,160,160)
from networks.Mednext.MedNextV1 import MedNeXt1 #(48,96,96)
from networks.Mednext.CNN3D import CNN3D#(48,96,96)
from networks.Mednext.MedCNN import MedCNN
from networks.Mednext.MSCNN3D import MSCNN3D
from networks.nSwinunetr.nSwinunetr import nSwinUNETR
from monai.transforms import AsDiscrete
from monai.metrics import DiceMetric
from monai.losses import DiceCELoss,DiceFocalLoss
from monai.inferers import sliding_window_inference
from monai.data import CacheDataset, DataLoader, decollate_batch,PersistentDataset
from torch.utils.tensorboard import SummaryWriter
from openpyxl import load_workbook
from datasets.load_datasets_transforms import data_transforms,dataset, RandomGenerator,Generator,patch_extraction,patch_reconstruction,roi_extraction
from torchvision import transforms
import random
from torchmetrics import Accuracy
from sklearn.utils.class_weight import compute_class_weight
import warnings
import openpyxl
from collections import defaultdict
import torch.nn.functional as F
from torch.autograd import Variable
from torch.cuda.amp import autocast, GradScaler
# 忽略所有警告
warnings.filterwarnings("ignore")

def config():
    parser = argparse.ArgumentParser(description='hyperparameters for medical image segmentation and classification')
    parser.add_argument('--dataset', type=str, default='301', help='Datasets: {feta, flare, amos}, Fyi: You can add your dataset here')
    parser.add_argument('--train_list', type=str, default='./data/train4.txt',help='Root folder')
    parser.add_argument('--val_list', type=str, default='./data/val4.txt', help='Root folder')
    parser.add_argument('--list_dir', type=str, default='./data/',help='Root folder')
    parser.add_argument('--seed', type=int,default=1234, help='random seed')
    #MedNeXt CNN3D+(32,80,64) MSCNN3D (48,96,96)
    parser.add_argument('--output', type=str, default='./output/MSCNN3D/0620e1/', help='Output folder for both tensorboard and the best model')
    parser.add_argument('--image_save', type=str, default='./output/MSCNN3D/0620e1', help='images')
    parser.add_argument('--network', type=str, default='MSCNN3D', help='Network models: {TransBTS, nnFormer, UNETR, SwinUNETR, 3DUXNET}')
    parser.add_argument('--supervision', type=bool,default=False, help='supervision')#True False
    parser.add_argument('--mode', type=str, default='train', help='Training or testing mode')
    parser.add_argument('--pretrain', default="True", help='Have pretrained weights or not')
    parser.add_argument('--pretrained_weights', default='', help='Path of pretrained weights')
    parser.add_argument('--patch', type=int, default=(32,160,160), help='Batch size for subject input')#(128,160,160) (80,128,128)
    parser.add_argument('--num_classes', type=int, default=4, help='Number of classes')
    parser.add_argument('--batch_size', type=int, default=7, help='Batch size for subject input')
    parser.add_argument('--crop_sample', type=int, default=1, help='Number of cropped sub-volumes for each subject')
    parser.add_argument('--lr', type=float, default=3e-4, help='Learning rate for training')
    parser.add_argument('--optim', type=str, default='AdamW', help='Optimizer types: Adam / AdamW')
    parser.add_argument('--max_iter', type=int, default=8000, help='Maximum iteration steps for training')
    parser.add_argument('--eval_step', type=int, default=250, help='Per steps to perform validation')#4000 1e8
    #(32,96,96) (32,128,128)  (32,160,160)
    ## Efficiency hyperparameters
    parser.add_argument('--gpu', type=str, default='0', help='your GPU number')
    parser.add_argument('--cache_rate', type=float, default=0.4, help='Cache rate to cache your dataset into GPUs')#0.1  缓存数据占总数的百分比!!
    parser.add_argument('--num_workers', type=int, default=6, help='Number of workers')

    args = parser.parse_args()
    # with open('./Yaml/3DX-Unet1.yaml', 'w') as f:
    #     yaml.dump(args.__dict__, f, indent=2)
    return args

def read_types():
    path = "./data/subtype_data_20240413.xlsx"  # "./subtype_data_20240413.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=14).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1

    path = "./data/Clinical_and_Other_Features_20230420.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=13).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1
    return type_dict

def seed_everything(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    #torch.backends.cudnn.benchmark = True / False

class FocalLoss(nn.Module):
    r"""
        This criterion is a implemenation of Focal Loss, which is proposed in
        Focal Loss for Dense Object Detection.

            Loss(x, class) = - \alpha (1-softmax(x)[class])^gamma \log(softmax(x)[class])

        The losses are averaged across observations for each minibatch.

        Args:
            alpha(1D Tensor, Variable) : the scalar factor for this criterion
            gamma(float, double) : gamma > 0; reduces the relative loss for well-classiﬁed examples (p > .5),
                                   putting more focus on hard, misclassiﬁed examples
            size_average(bool): By default, the losses are averaged over observations for each minibatch.
                                However, if the field size_average is set to False, the losses are
                                instead summed for each minibatch.


    """
    def __init__(self, class_num=6, alpha=None, gamma=2, size_average=True):
        super(FocalLoss, self).__init__()
        if alpha is None:
            self.alpha = Variable(torch.ones(class_num, 1))
        else:
            if isinstance(alpha, Variable):
                self.alpha = alpha
            else:
                self.alpha = Variable(alpha)
        self.gamma = gamma
        self.class_num = class_num
        self.size_average = size_average

    def forward(self, inputs, targets):
        N = inputs.size(0)
        C = inputs.size(1)
        P = F.softmax(inputs)

        class_mask = inputs.data.new(N, C).fill_(0)
        class_mask = Variable(class_mask)
        ids = targets.view(-1, 1)
        class_mask.scatter_(1, ids.data, 1.)
        #print(class_mask)


        if inputs.is_cuda and not self.alpha.is_cuda:
            self.alpha = self.alpha.cuda()
        alpha = self.alpha[ids.data.view(-1)]

        probs = (P*class_mask).sum(1).view(-1,1)

        log_p = probs.log()
        #print('probs size= {}'.format(probs.size()))
        #print(probs)

        batch_loss = -alpha*(torch.pow((1-probs), self.gamma))*log_p
        #print('-----bacth_loss------')
        #print(batch_loss)


        if self.size_average:
            loss = batch_loss.mean()
        else:
            loss = batch_loss.sum()
        return loss

class MulticlassHingeLoss(nn.Module):
    def __init__(self, margin=1.0):
        super(MulticlassHingeLoss, self).__init__()
        self.margin = margin

    def forward(self, inputs, targets):
        num_classes = inputs.size(1)
        one_hot_targets = nn.functional.one_hot(targets, num_classes=num_classes).float()
        correct_class_scores = torch.sum(inputs * one_hot_targets, dim=1)
        margins = torch.max(inputs - correct_class_scores[:, None] + self.margin, torch.zeros_like(inputs))
        loss = torch.sum(margins * one_hot_targets) / inputs.size(0)
        return loss

def multiclass_hinge_loss(logits, labels, margin=1.0):
    num_classes = logits.shape[-1]
    one_hot_labels = torch.zeros_like(logits).scatter_(1, labels.unsqueeze(1), 1)
    logits = logits - one_hot_labels * margin
    logits = torch.clamp(logits, min=0)
    loss = 1 - logits.sum(dim=1)
    return loss.mean()

if __name__ == '__main__':
    # seed_everything(seed=42)
    set_determinism(seed=1234)
    types=read_types()
    args=config() #the first running should excute this code
    # device="cpu"
    # os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu
    # print('Used GPU: {}'.format(args.gpu))
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print('Used GPU: {}'.format(device))
    train_transforms, val_transforms = data_transforms(args)
    
    # db_train = dataset(list_dir=args.list_dir, split="train", num_classes=args.num_classes,
    #                            transform=train_transforms)#0225
    db_train = dataset(list_dir=args.train_list, split="train", num_classes=args.num_classes,
                       transform=transforms.Compose([RandomGenerator(output_size=args.patch, mode='train')]))#0226
    print("The length of train set is: {}".format(len(db_train)))
    db_val = dataset(list_dir=args.val_list, split="val", num_classes=args.num_classes,
                               transform=transforms.Compose([Generator(output_size=args.patch, mode = 'val')]))
    print("The length of val set is: {}".format(len(db_val)))


    def get_filename(filename):
        return filename.split('/')[-2]

    class_counts = defaultdict(int)
    class_mapping = {}  # 用于映射类别标识符到类别的字典，比如 {'IO123': 1, 'IO113': 4, ...}
    max_count = 0
    for img_path in db_train.sample_list:
        class_id = get_filename(img_path)#img_path.split('/')[-2]  # 假设类别ID是文件路径倒数第二个斜杠后的部分
        values = types[class_id]
        class_mapping[class_id] =values
        class_counts[values] += 1
        max_count = max(max_count, class_counts[values])  # 更新最大样本数
    balanced_data = []
    for class_id, count in class_counts.items():
        original_samples = [filename for filename in db_train.sample_list if types[get_filename(filename)] == class_id]
        aa = (max_count - count) // len(original_samples)
        samples_to_add = original_samples * aa
        if (max_count - count) % len(original_samples) > 0:
            samples_to_add.extend(random.sample(original_samples, (max_count - count) % len(original_samples)))
        balanced_data.extend(samples_to_add)
    # db_train.sample_list.extend(balanced_data)
    values_list=[]
    for i in range(len(db_train.sample_list)):
        img_path = db_train.sample_list[i].strip('\n')
        ID = img_path.split('/')[-2]
        values =types[ID]# [types[key] for key in ID if key in types]
        values_list.append(int(values))
    ID_value=np.asarray(values_list).astype(int)
    ID_value=np.sort(ID_value)
    class_weights = compute_class_weight('balanced', classes=np.unique(ID_value), y=ID_value)  # 假设有1到5共5个类别
    class_weights = np.insert(class_weights, 0, 5)
    class_weights = torch.tensor(class_weights, dtype=torch.float).to(device)

    def worker_init_fn(worker_id):
        random.seed(args.seed + worker_id)

    train_loader = DataLoader(db_train, batch_size=args.batch_size, shuffle=True, num_workers=6, pin_memory=True,
                             worker_init_fn=worker_init_fn)
    val_loader = DataLoader(db_val, batch_size=1, shuffle=True, num_workers=6, pin_memory=True,
                             worker_init_fn=worker_init_fn)

    ## Load Networks
    # Initialize the model
    if args.network =='unet':##(1,1,96,96,96)
        model = unet().to(device)
    elif args.network == 'SegResNet':
        model = SegResNet().to(device)
    elif args.network == 'BasicUNet':
        model = BasicUNet().to(device)
    elif args.network == "MedNeXt":  ##(160,160,96)
        model = MedNeXt(
       in_channels = 3,
        n_channels = 32,
        n_classes = args.num_classes,
        exp_r=[2,3,4,4,4,4,4,3,2],
        kernel_size=5,
        deep_supervision=args.supervision,
        do_res=True,
        do_res_up_down = True,
        block_counts = [3,4,4,4,4,4,4,4,3],
        checkpoint_style = 'outside_block'
       # model = MedNeXt1(
       #  in_channels=3,
       #  n_channels=32,
       #  n_classes=args.num_classes,
       #  exp_r=[3, 4, 8, 8, 8, 8, 8, 4, 3],
       #  kernel_size=3,
       #  deep_supervision=args.supervision,
       #  do_res=True,
       #  do_res_up_down=True,
       #  block_counts=[3, 4, 8, 8, 8, 8, 8, 4, 3],
       #  checkpoint_style='outside_block'
        ).to(device)
    elif args.network == "nSwinUNETR":  # (160,160,96)
        model = nSwinUNETR(
            img_size=args.patch,
            in_channels=1,
            out_channels=args.num_classes,
            feature_size=48,
            drop_rate=0.0,
            attn_drop_rate=0.0,
            dropout_path_rate=0.0,
            use_checkpoint=True,
            deep_supervision=args.supervision,
        ).to(device)
    elif args.network == "MSCNN3D":  ##(160,160,96)
        model = MSCNN3D(args.num_classes).to(device)
    elif args.network == "CNN3D":  ##(160,160,96)
        model = CNN3D(
            in_channels=3,
            out_channels=64,
            num_classes=args.num_classes
        ).to(device)
    print('Chosen Network Architecture: {}'.format(args.network))
    
    if args.pretrain == 'False':# True False
        print('Pretrained weight is found! Start to load weight from')
        modd = "best.pth"
        model.load_state_dict(torch.load(os.path.join(args.output, modd)))
    loss1 = MulticlassHingeLoss(margin=2)  # class_weight
    loss2 = nn.CrossEntropyLoss()
    print('Loss for training: {}'.format('DiceCELoss'))
    if args.optim == 'AdamW':
        optimizer = torch.optim.AdamW(model.parameters(), lr=args.lr)
    elif args.optim == 'Adam':
        optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)
    print('Optimizer for training: {}, learning rate: {}'.format(args.optim, args.lr))
    # scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, 'min', factor=0.9, patience=1000)
    scaler = GradScaler() if device.type == 'cuda' else None

    root_dir = os.path.join(args.output)
    if os.path.exists(root_dir) == False:
        os.makedirs(root_dir)

    t_dir = os.path.join(root_dir, 'tensorboard')
    if os.path.exists(t_dir) == False:
        os.makedirs(t_dir)
    writer = SummaryWriter(log_dir=t_dir)

    def validation(epoch_iterator_val):
        model.eval()
        dis_vals=[]
        pred=[]
        gd=[]
        correct = 0
        total = 0
        with torch.no_grad():
            for step, batch in enumerate(epoch_iterator_val):
                torch.cuda.empty_cache()
                # type = types[batch["ID"]].long()
                # values_list = [types[key] for key in batch["ID"] if key in types]
                t = batch[0]["label"]
                catageries = torch.amax(t, dim=(1, 2, 3, 4))
                catageries[(catageries >= 0.5) & (catageries <= 2.5)] = 1
                catageries[(catageries >= 2.5) & (catageries <= 4.5)] = 2
                catageries[(catageries >= 4.5)] = 3
                catageries = torch.Tensor(catageries).long().to(device)
                patch=batch[0]["image"]
                # for patch in patchs:
                predict_type = model(patch.to(device))
                _, predicted = torch.max(predict_type.data, 1)
                total += catageries.size(0)
                correct += (predicted == catageries).sum().item()
        mean_dis_val = correct / total
        #             _, predicted_classes = torch.max(predict_type, dim=1)
        #             pred.append(predicted_classes.cpu().detach())
        #             gd.append(catageries)
        # print("number of diagnosis:"+str(len(gd)))
        # mean_dis_val = accuracy_metric(torch.tensor(pred),torch.tensor(gd))
        writer.add_scalar('Validation Segmentation Loss', mean_dis_val, global_step)
        return mean_dis_val

    def train(global_step, train_loader, dice_val_best,dis_val_best, global_step_best):
        model.train()
        epoch_loss = 0
        epoch_iterator = tqdm(train_loader, desc="Training (X / X Steps) (loss=X.X)", dynamic_ncols=True)
        for step, batch in enumerate(epoch_iterator):
            step += 1
            # values_list = [types[key] for key in batch[1]["ID"] if key in types]
            # catageries=torch.Tensor(values_list).long()
            # catageries=catageries*batch["flag"]
            t=batch[0]["label"]
            catageries = torch.amax(t, dim=(1, 2, 3, 4))
            catageries[(catageries >= 0.5) & (catageries <= 2.5)] = 1
            catageries[(catageries >= 2.5) & (catageries <= 4.5)] = 2
            catageries[(catageries >= 4.5)] = 3
            # image_batch, label_batch = batch['image'], batch['label']
            x,z = (batch[0]["image"].to(device), catageries.to(device))
            with autocast():  # 使用autocast上下文管理器，自动将接下来的操作转换为混合精度
                predict_type = model(x)
                # loss = loss1(predict_type, z)+loss2(predict_type, z)*0.5
                loss = loss2(predict_type, z)

            scaler.scale(loss).backward()  # 使用scaler来缩放损失并进行反向传播
            scaler.step(optimizer)  # 更新优化器的参数
            scaler.update()  # 更新scaler的状态
            optimizer.zero_grad()
            # predict_type = model(x)
            # loss=loss2(predict_type,z)
            # loss.backward()
            # epoch_loss += loss.item()
            # optimizer.step()
            # optimizer.zero_grad()
            epoch_loss += loss.item()
            epoch_iterator.set_description("Training (%d / %d Steps) (loss=%2.5f)" % (global_step, max_iterations, loss)
            )

            if (
                    global_step % eval_num == 0 and global_step != 0
            ) or global_step == max_iterations:
                epoch_iterator_val = tqdm(
                    val_loader, desc="Validate (X / X Steps) (dice=X.X)", dynamic_ncols=True
                )
                dis_val = validation(epoch_iterator_val)
                epoch_loss /= step
                epoch_loss_values.append(epoch_loss)
                metric_values.append(dis_val)
                if dis_val >= dis_val_best:
                    dis_val_best = dis_val
                    global_step_best = global_step
                    # modd = str(datetime.datetime.now().strftime('%m_%d_%H')) + ".pth"
                    modd="best.pth"
                    torch.save(model.state_dict(), os.path.join(root_dir, modd))
                    print(
                        "Model Was Saved ! Current Best Avg. acc: {} Current Avg. acc: {}".format(
                            dis_val_best, dis_val))
                    # scheduler.step(dice_val)
                else:
                    print("Model Was Not Saved ! Current Best Avg. acc: {} Current Avg. acc: {}".format(
                            dis_val_best, dis_val))
                    # scheduler.step(dice_val)
                # del epoch_iterator_val  # 清除验证迭代器引用
                # torch.cuda.empty_cache()  # 尝试释放缓存的无用内存
            writer.add_scalar('Training Segmentation Loss', loss.data, global_step)
            global_step += 1
        return global_step, dice_val_best,dis_val_best, global_step_best

    max_iterations = args.max_iter
    print('Maximum Iterations for training: {}'.format(str(args.max_iter)))
    eval_num = args.eval_step
    post_label = AsDiscrete(to_onehot=args.num_classes)
    post_pred = AsDiscrete(argmax=True, threshold=0.5,to_onehot=args.num_classes)
    dice_metric = DiceMetric(include_background=False, reduction="none", get_not_nans=False)  # 去除背景项目
    accuracy_metric = Accuracy(task="multiclass",num_classes=12)
    dice_val_best=0.0
    dis_val_best = 0.0
    global_step_best = 0
    epoch_loss_values = []
    metric_values = []
    global_step=0
    while global_step < max_iterations:
        global_step, dice_val_best,dis_val_best, global_step_best = train(
            global_step, train_loader, dice_val_best,dis_val_best, global_step_best
        )