from batchgenerators.utilities.file_and_folder_operations import *
import shutil
from pathlib import Path
from nnunetv2.dataset_conversion.generate_dataset_json import generate_dataset_json
from nnunetv2.paths import nnUNet_raw,nnUNet_preprocessed
from batchgenerators.utilities.file_and_folder_operations import save_json, join,load_json
import random
import numpy as np
import SimpleITK as sitk
import numpy as np
import nibabel as nib
import copy
import json
import openpyxl

def read_types():
    path = "./subtype_data_20240413.xlsx"  # "./subtype_data_20240413.xlsx"
    look_up_table_row_start = 3
    look_up_table_row_number = 400  # 根据实际行数调整
    type_dict = {}
    # wb = openpyxl.load_workbook(path)
    # sheet = wb[wb.sheetnames[0]]
    # for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
    #     ID = sheet.cell(row=i, column=1).value
    #     # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
    #     sub_data1 = sheet.cell(row=i, column=14).value
    #     if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
    #         type_dict[ID] = sub_data1

    path = "./Clinical_and_Other_Features_20230420.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb[wb.sheetnames[0]]
    for i in range(look_up_table_row_start, look_up_table_row_start + look_up_table_row_number):
        ID = sheet.cell(row=i, column=1).value
        # 检查单元格值是否为有效数据，如果不是（例如：np.nan），则跳过此次循环
        sub_data1 = sheet.cell(row=i, column=13).value
        if isinstance(sub_data1, (str, int, float)) and not np.isnan(sub_data1):
            type_dict[ID] = sub_data1
    return type_dict

def make_out_dirs(dataset_id: int, task_name="Breast"):
    dataset_name = f"Dataset{dataset_id:03d}_{task_name}"

    out_dir = Path(nnUNet_raw.replace('"', "")) / dataset_name
    train_img_dir = out_dir / "imagesTr"
    train_labels_dir = out_dir / "labelsTr"
    test_img_dir = out_dir / "imagesTs"
    test_labels_dir = out_dir / "labelsTs"
    os.makedirs(out_dir, exist_ok=True)

    os.makedirs(train_img_dir, exist_ok=True)
    os.makedirs(train_labels_dir, exist_ok=True)
    os.makedirs(test_img_dir, exist_ok=True)
    os.makedirs(test_labels_dir, exist_ok=True)

    return out_dir, train_img_dir, train_labels_dir, test_img_dir,test_labels_dir

def create_split(labelsTr_folder: str, seed: int = 1234) -> List[dict[str, List]]:
    map = os.path.join(nnUNet_preprocessed, "Dataset301_Breast", "map.json")
    fmap = open(map)
    data_dict = {}
    for line in fmap.readlines():
        if line.strip():  # 确保这一行不是空白
            obj = json.loads(line)
            # 每行应该只有一个键值对，所以我们取这个对象的第一个键值对
            key, value = list(obj.items())[0]
            data_dict[key] = value

    ope = "../breast_list/open.txt"
    f = open(ope)  # train
    open_cases=[]
    for line in f.readlines():
        # train_list.append(line.split('\n')[0])
        known_path=line.split('\n')[0]
        for key, value in data_dict.items():
            if value == known_path:
                open_cases.append(key)
    val_cases=[]
    splits = []
    for fold in range(5):
        train="../breast_list/train"+str(fold)+".txt"
        f = open(train)  # train
        train_list = []  # train
        train_cases=[]
        for line in f.readlines():
            # train_list.append(line.split('\n')[0])
            known_path=line.split('\n')[0]
            for key, value in data_dict.items():
                if value == known_path:
                    train_cases.append(key)
        train_cases=train_cases+open_cases
        train_list.sort()
        val_cases=[]
        val="../breast_list/val"+str(fold)+".txt"
        ff = open(val)  # train
        val_list = []  # train
        for line in ff.readlines():
            # val_list.append(line.split('\n')[0])
            known_path = line.split('\n')[0]
            for key, value in data_dict.items():
                if value == known_path:
                    val_cases.append(key)
        val_list.sort()
        splits.append({'train': train_cases, 'val': val_cases})
    return splits

def copy_files(src_data_folder: Path, train_dir: Path, labelTr_dir: Path, test_dir: Path,labelTs_dir: Path):
    """Copy files from the ACDC dataset to the nnUNet dataset folder. Returns the number of training cases."""
    path_list = []
    for root, dirs, files in os.walk(src_data_folder, topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path and "normal" not in path:
                path_list.append(path)

    #open dataset
    types = read_types()
    f=open("../breast_list/open.txt","w")
    for root, dirs, files in os.walk(src_data_folder.replace("set1","set2"), topdown=False):
        for file in files:
            path = os.path.join(root, file)
            if "0.nii.gz" in path and "normal" not in path:
                ID=path.split('/')[-2]
                value=types[ID]
                if value>2 and value<5:
                    readd = sitk.ReadImage(path.replace("0.nii.gz", "1.nii.gz"), sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
                    lab = sitk.GetArrayFromImage(readd)
                    if lab.max()<1:
                        print(path)
                        continue
                    f.writelines(path + "\n")
                    path_list.append(path)
    f.close()

    pathh=os.path.join(nnUNet_preprocessed,"Dataset301_Breast")
    os.makedirs(pathh, exist_ok=True)
    map = os.path.join(nnUNet_preprocessed, "Dataset301_Breast", "map.json")
    mapObject = open(map, 'a', encoding='utf-8')
    mapObject.seek(0)  #
    mapObject.truncate()  # clear content

    num_cases = 0
    num_training_cases=0
    num_test_cases = 0
    ii = 0
    for path_str in path_list:
        file = Path(path_str)
        if file.name == "0.nii.gz":

            new_data = {f"Breast_{str(num_cases).zfill(4)}": path_str}
            js = json.dumps(new_data, ensure_ascii=False)
            mapObject.write(js + '\n')
            # num_cases += 1
            # continue

            # We split the stem and append _0000 to the patient part.
            se0output = train_dir / f"Breast_{str(num_cases).zfill(4)}_0000.nii.gz"
            read0 = sitk.ReadImage(file, sitk.sitkInt16)
            img_array00 = sitk.GetArrayFromImage(read0)
            image_size = img_array00.shape
            if "set2" in path_str:
                aa = img_array00.max()
                t = aa / 1024
                img_array00 = img_array00 //t

            label_path = path_str.replace("0.nii.gz", "1.nii.gz")
            se1output = labelTr_dir / f"Breast_{str(num_cases).zfill(4)}.nii.gz"
            read1 = sitk.ReadImage(label_path, sitk.sitkInt16)  # 使用sitk重新保存，这样占用内存小很多
            label = sitk.GetArrayFromImage(read1)

            left_mask = copy.deepcopy(label)
            right_mask = copy.deepcopy(label)
            left_mask[:, :, image_size[1] // 2:] = 0
            right_mask[:, :, :image_size[1] // 2] = 0
            if np.sum(left_mask) >= np.sum(right_mask):
                img_array11 = left_mask
            else:
                img_array11 = right_mask

            index = np.nonzero(img_array11)
            index = np.transpose(index)
            z_min, y_min, x_min = index.min(axis=0)
            z_max, y_max, x_max = index.max(axis=0)
            img_array = img_array00[z_min:z_max + 1, y_min:y_max + 1, x_min:x_max + 1]
            label_array = img_array11[z_min:z_max + 1, y_min:y_max + 1, x_min:x_max + 1]
            #step1
            # label_array[(label_array >= 0.5) & (label_array <= 4.5)] = 1
            # label_array[label_array > 4.5] = 2

            #step2
            if img_array11.max()>4.5:
                continue
            label_array[(label_array >= 0.5) & (label_array <= 2.5)] = 1
            label_array[(label_array >= 2.5) & (label_array <= 4.5)] = 2


            out0 = sitk.GetImageFromArray(img_array)
            sitk.WriteImage(out0, se0output)
            out1 = sitk.GetImageFromArray(label_array)
            sitk.WriteImage(out1, se1output)
            num_cases += 1
        else:
            print(path_str)
        ii = ii + 1
        if ii % 10 == 0:
            print('numbers:', ii)
    mapObject.close()
    return num_cases,num_test_cases

def convert_Breast(src_data_folder: str, dataset_id=301):
    out_dir, train_img_dir, train_labels_dir, test_img_dir,test_labels_dir = make_out_dirs(dataset_id=dataset_id)
    num_training_cases,num_test_cases= copy_files(src_data_folder, train_img_dir, train_labels_dir, test_img_dir,test_labels_dir)

    generate_dataset_json(
        str(out_dir),
        channel_names={
            0: "MRI",
        },
        labels={
            "background": 0,
            "lesion1": 1,
            "lesion2": 2,
        },
        file_ending=".nii.gz",
        num_training_cases=num_training_cases,
    )

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-i",
        "--input_folder",
        type=str,
        help="The downloaded Breast dataset dir. Should contain extracted 'training' and 'testing' folders.",
    )
    parser.add_argument(
        "-d", "--dataset_id", required=False, type=int, default=301, help="nnU-Net Dataset ID, default: 301"
    )
    args = parser.parse_args()
    print("Converting...")
    input_folder="/media/bit301/data/yml/data/xy/breast/set1"
    dataset_id=301
    # convert_Breast(input_folder, dataset_id)

    #five_cross_validation
    dataset_name = f"Dataset{args.dataset_id:03d}_{'Breast'}"
    labelsTr = join(nnUNet_raw, dataset_name, 'labelsTr')
    preprocessed_folder = join(nnUNet_preprocessed, dataset_name)
    maybe_mkdir_p(preprocessed_folder)
    split = create_split(labelsTr)
    save_json(split, join(preprocessed_folder, 'splits_final.json'), sort_keys=False)
    print("Done!")